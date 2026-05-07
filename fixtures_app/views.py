#IO:crear archivos en memoria PDF y EXCEL
import io
from datetime import timedelta, date

#openpyxl: crear archivos .xlsx y estilos 
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

#reportlab: genera pdf dinamicos
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

#Renderizar HTML, redireccionar y generar erro si un objeto no existe
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse

#Imports propios
from .models import Equipo, Torneo, Jornada, Partido
from .utils import generar_fixture


# ── Lista de equipos ──────────────────────────────────────────────
def lista_equipos(request):
    equipos = Equipo.objects.all()
    return render(request, 'fixtures_app/equipos.html', {'equipos': equipos})
""" 

# ── Crear torneo + seleccionar equipos ───────────────────────────
def crear_torneo(request):
    equipos = Equipo.objects.all()

    if request.method == 'POST':
        nombre       = request.POST.get('nombre')
        fecha_inicio = request.POST.get('fecha_inicio')
        ids_equipos  = request.POST.getlist('equipos')  # lista de IDs

        if len(ids_equipos) < 4:
            error = 'Selecciona al menos 4 equipos.'
            return render(request, 'fixtures_app/crear_torneo.html',
                          {'equipos': equipos, 'error': error})

        torneo = Torneo.objects.create(nombre=nombre, fecha_inicio=fecha_inicio)
        equipos_sel = Equipo.objects.filter(id__in=ids_equipos)
        torneo.equipos.set(equipos_sel)

        # Generar fixture
        jornadas_data = generar_fixture(list(equipos_sel))
        fecha_actual  = date.fromisoformat(fecha_inicio)

        for num, partidos in enumerate(jornadas_data, start=1):
            jornada = Jornada.objects.create(torneo=torneo, numero=num)
            for local, visitante in partidos:
                Partido.objects.create(
                    jornada=jornada,
                    equipo_local=local,
                    equipo_visitante=visitante,
                    fecha=fecha_actual,
                )
            fecha_actual += timedelta(weeks=1)

        return redirect('ver_fixture', torneo_id=torneo.id)

    return render(request, 'fixtures_app/crear_torneo.html', {'equipos': equipos})


# ── Ver fixture en pantalla ───────────────────────────────────────
def ver_fixture(request, torneo_id):
    torneo   = get_object_or_404(Torneo, id=torneo_id)
    jornadas = torneo.jornadas.prefetch_related('partidos__equipo_local',
                                                'partidos__equipo_visitante')
    return render(request, 'fixtures_app/fixture.html',
                  {'torneo': torneo, 'jornadas': jornadas})


# ── Exportar PDF ──────────────────────────────────────────────────
def exportar_pdf(request, torneo_id):
    torneo   = get_object_or_404(Torneo, id=torneo_id)
    jornadas = torneo.jornadas.prefetch_related('partidos__equipo_local',
                                                'partidos__equipo_visitante')

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=A4)
    estilos = getSampleStyleSheet()
    elementos = []

    titulo = Paragraph(f"<b>Fixture: {torneo.nombre}</b>", estilos['Title'])
    elementos.append(titulo)
    elementos.append(Spacer(1, 20))

    for jornada in jornadas:
        elementos.append(Paragraph(f"<b>Jornada {jornada.numero}</b>", estilos['Heading2']))
        datos = [['Local', 'VS', 'Visitante', 'Fecha']]
        for p in jornada.partidos.all():
            datos.append([
                p.equipo_local.nombre,
                'vs',
                p.equipo_visitante.nombre,
                str(p.fecha) if p.fecha else '-',
            ])

        tabla = Table(datos, colWidths=[160, 30, 160, 100])
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a3a5c')),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#eef2f7')]),
            ('GRID',       (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE',   (0, 0), (-1, -1), 10),
        ]))
        elementos.append(tabla)
        elementos.append(Spacer(1, 16))

    doc.build(elementos)
    buffer.seek(0)
    resp = HttpResponse(buffer, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="fixture_{torneo.nombre}.pdf"'
    return resp


# ── Exportar Excel ────────────────────────────────────────────────
def exportar_excel(request, torneo_id):
    torneo   = get_object_or_404(Torneo, id=torneo_id)
    jornadas = torneo.jornadas.prefetch_related('partidos__equipo_local',
                                                'partidos__equipo_visitante')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = torneo.nombre[:31]

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1a3a5c')
    center      = Alignment(horizontal='center')

    fila = 1
    for jornada in jornadas:
        # Encabezado jornada
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=4)
        celda = ws.cell(row=fila, column=1, value=f'Jornada {jornada.numero}')
        celda.font = Font(bold=True, color='FFFFFF')
        celda.fill = PatternFill('solid', fgColor='2d6a9f')
        celda.alignment = center
        fila += 1

        # Cabeceras columnas
        for col, texto in enumerate(['Local', 'VS', 'Visitante', 'Fecha'], start=1):
            c = ws.cell(row=fila, column=col, value=texto)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
        fila += 1

        for i, p in enumerate(jornada.partidos.all()):
            fill_color = 'FFFFFF' if i % 2 == 0 else 'EEF2F7'
            fila_fill  = PatternFill('solid', fgColor=fill_color)
            for col, val in enumerate(
                [p.equipo_local.nombre, 'vs', p.equipo_visitante.nombre,
                 str(p.fecha) if p.fecha else '-'], start=1
            ):
                c = ws.cell(row=fila, column=col, value=val)
                c.fill = fila_fill
                c.alignment = center
            fila += 1
        fila += 1  # espacio entre jornadas

    for col in ws.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    resp = HttpResponse(buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="fixture_{torneo.nombre}.xlsx"'
    return resp """